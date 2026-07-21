"""
SEM süre güncelleme aracı.

Çalıştır:
    python scripts/apply_updates.py

Menü:
  1 → sporcu_yildizlar.xlsx ile karşılaştır, sure_guncelleme.xlsx'i yenile
  2 → sure_guncelleme.xlsx'te "E" olanları SEM_kontrol_liste.xlsx'e aktar + paneli güncelle + GitHub'a push
"""
import sys, os, re, unicodedata, shutil, subprocess
from datetime import datetime
sys.stdout.reconfigure(encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT       = os.path.dirname(SCRIPT_DIR)
SEM_ENTRY  = os.path.join(ROOT, "SEM_kontrol_liste.xlsx")
UPDATE_XL  = os.path.join(ROOT, "data", "sure_guncelleme.xlsx")
BACKUP_DIR = os.path.join(ROOT, "backups")

# ─────────────────────────────────────────────────────────────────────────────
def ask(prompt):
    try:
        return input(prompt).strip()
    except (EOFError, KeyboardInterrupt):
        print("\nİptal edildi.")
        sys.exit(0)

def normalize(name):
    if not name: return ""
    name = str(name).upper().strip()
    tr = str.maketrans("ÇĞİIÖŞÜçğışöşü", "CGIIOSUcgiossu")
    name = name.translate(tr)
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", name)

# ─────────────────────────────────────────────────────────────────────────────
def run_compare():
    """compare_times.py çalıştır — sure_guncelleme.xlsx yenilenir."""
    print("\nKarşılaştırma başlıyor...")
    result = subprocess.run(
        [sys.executable, os.path.join(SCRIPT_DIR, "compare_times.py")],
        capture_output=False, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        print("HATA: karşılaştırma başarısız.")
        sys.exit(1)
    print("\nKarşılaştırma tamamlandı.")
    print(f"Sonuçlar: {UPDATE_XL}")
    print("Dosyayı aç, 'E' onayladıklarına yaz, kaydet.")
    print("Sonra tekrar çalıştır ve seçenek 2'yi seç.")

# ─────────────────────────────────────────────────────────────────────────────
def run_apply():
    """sure_guncelleme.xlsx'teki E olanları SEM listesine uygula + push."""
    import openpyxl

    if not os.path.exists(UPDATE_XL):
        print(f"HATA: {UPDATE_XL} bulunamadı.")
        print("Önce seçenek 1 ile karşılaştırma yapın.")
        sys.exit(1)

    # ── Onaylananları oku ────────────────────────────────────────────────────
    wb_u = openpyxl.load_workbook(UPDATE_XL, data_only=True)
    ws_u = wb_u["Güncellenecekler"]

    approved = []  # [(norm_name, yb, brans, yeni_sure, orijinal_ad)]
    count_s = count_h = 0
    for r in range(2, ws_u.max_row + 1):
        name  = ws_u.cell(r, 2).value
        if not name:
            continue
        yb    = str(ws_u.cell(r, 3).value or "").strip()
        brans = str(ws_u.cell(r, 7).value or "").strip()
        yeni  = str(ws_u.cell(r, 9).value or "").strip()
        onay  = str(ws_u.cell(r, 11).value or "").strip().upper()
        if onay == "E":
            approved.append((normalize(str(name)), yb, brans, yeni, str(name)))
        elif onay == "H":
            count_h += 1
        else:
            # "S", "SONRA İNCELEYECEĞİM", boş, vb. → bekliyor
            count_s += 1

    print(f"\n  Onaylanan (E): {len(approved)}")
    print(f"  Bekleyen  (S): {count_s}")
    print(f"  Reddedilen(H): {count_h}")

    if not approved:
        print("\nGüncellenecek 'E' kayıt yok. Çıkılıyor.")
        sys.exit(0)

    print(f"\nGüncellenecekler:")
    for _, yb, brans, yeni, orijinal in approved:
        print(f"  {orijinal:30s}  YB={yb}  {brans:20s}  → {yeni}")

    onay = ask(f"\n{len(approved)} satır SEM_kontrol_liste.xlsx'e yazılacak. Devam? (e/H): ")
    if onay.lower() != "e":
        print("İptal edildi.")
        sys.exit(0)

    # ── Backup ───────────────────────────────────────────────────────────────
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"SEM_kontrol_liste_{ts}.xlsx")
    shutil.copy2(SEM_ENTRY, backup_path)
    print(f"Backup alındı: {backup_path}")

    # ── SEM_kontrol_liste güncelle ───────────────────────────────────────────
    wb_sem = openpyxl.load_workbook(SEM_ENTRY)
    ws_sem = wb_sem["SEM Kontrol Listesi"]

    row_index = {}
    for r in range(2, ws_sem.max_row + 1):
        name = ws_sem.cell(r, 4).value
        yb   = str(ws_sem.cell(r, 5).value or "").strip()
        if name:
            row_index[(normalize(str(name)), yb)] = r

    updated   = 0
    not_found = []
    for norm_name, yb, brans, yeni, orijinal in approved:
        key = (norm_name, yb)
        if key not in row_index:
            not_found.append((orijinal, yb, brans))
            continue
        r = row_index[key]
        done = False
        for i in range(4):
            brans_col = 7 + i * 2
            sure_col  = 8 + i * 2
            cell_brans = ws_sem.cell(r, brans_col).value
            if cell_brans and str(cell_brans).strip() == brans:
                old_val = ws_sem.cell(r, sure_col).value
                ws_sem.cell(r, sure_col).value = yeni
                print(f"  OK  {ws_sem.cell(r,4).value:30s}  {brans:20s}  {old_val} → {yeni}")
                updated += 1
                done = True
                break
        if not done:
            not_found.append((orijinal, yb, brans))

    wb_sem.save(SEM_ENTRY)
    print(f"\nGüncellendi: {updated}")
    if not_found:
        print("Bulunamayan:")
        for n, y, b in not_found:
            print(f"  {n}  YB={y}  {b}")

    # ── Panel verisini yenile ────────────────────────────────────────────────
    print("\npanel/data.js yenileniyor...")
    result = subprocess.run(
        [sys.executable, os.path.join(SCRIPT_DIR, "sem_generate.py")],
        cwd=ROOT, capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        print("HATA:", result.stderr)
        sys.exit(1)
    print("  OK — panel/data.js güncellendi")

    # ── Git commit + push ────────────────────────────────────────────────────
    print("\nGitHub'a gönderiliyor...")
    os.chdir(ROOT)

    def git(args, check=True):
        r = subprocess.run(["git"] + args, cwd=ROOT, capture_output=True, text=True, encoding="utf-8")
        if check and r.returncode not in (0, 1):
            print("git hata:", r.stderr)
        return r

    git(["add", "SEM_kontrol_liste.xlsx", "panel/data.js", "data/results.json"])
    msg = f"Sure guncelleme: {updated} brans ({len(approved)} onaylandi)"
    git(["commit", "-m", msg])
    git(["pull", "origin", "main", "-X", "ours", "--no-edit"], check=False)
    push = git(["push", "origin", "main"])
    if push.returncode == 0:
        print(f"  OK — GitHub'a push edildi")
        print(f"\nPanel birkaç dakika içinde güncellenir:")
        print("  https://gokaygunduz-gg.github.io/sem-samsun-2026/panel/index.html")
    else:
        print("Push hatası:", push.stderr)
        print("Manuel olarak push edin:")
        print("  git pull origin main -X ours --no-edit && git push origin main")

# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  SEM Süre Güncelleme Aracı")
    print("=" * 60)
    print()
    print("  1 → sporcu_yildizlar.xlsx ile karşılaştır")
    print("       (sure_guncelleme.xlsx yenilenir, onayınıza sunulur)")
    print()
    print("  2 → sure_guncelleme.xlsx'teki 'E' olanları aktar")
    print("       (SEM listesi güncellenir, panel yenilenir, GitHub'a push)")
    print()
    secim = ask("Seçiminiz (1/2): ")

    if secim == "1":
        run_compare()
    elif secim == "2":
        run_apply()
    else:
        print("Geçersiz seçim.")

if __name__ == "__main__":
    main()
