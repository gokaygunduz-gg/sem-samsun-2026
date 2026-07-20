"""
SEM giriş listesi okuyucu.
SEM_kontrol_liste.xlsx → sporcu listesi (dict listesi).

Sütun yapısı (1-indexed):
  1: S.     2: Şehir   3: Cinsiyet   4: Ad Soyad
  5: YB     6: TC No
  7: Yarış1  8: Süre1  9: Yarış2  10: Süre2
  11: Yarış3 12: Süre3 13: Yarış4 14: Süre4
  16: Kulüp
"""

import openpyxl
from sem_config import ENTRY_PATH, NO_50M_GROUPS, PROGRAM

# Giriş listesinde yalnızca program içindeki branşlara izin ver
KNOWN_EVENTS = {b for _, _, b in PROGRAM}

TR_MAP = str.maketrans({
    "ğ": "g", "Ğ": "G", "ı": "i", "İ": "i",
    "ş": "s", "Ş": "S", "ü": "u", "Ü": "U",
    "ö": "o", "Ö": "O", "ç": "c", "Ç": "C",
})

# Şehir yazım düzeltmeleri (Latin → Türkçe karakterli resmi ad)
CITY_FIXES: dict[str, str] = {
    "istanbul": "İstanbul",
    "izmir":    "İzmir",
}

# ---------------------------------------------------------------------------
# Sporcu adı kanonik tablosu: (norm(ham_ad), yb) → doğru görüntü adı
#
# Kural: norm() fonksiyonu İ/I/ı/i farklarını zaten ortadan kaldırır,
# bu yüzden her sporcu için TEK bir anahtar yeterlidir.
# Tüm varyantlar (ırmak/Irmak/İrmak) aynı norma düşer.
# ---------------------------------------------------------------------------
NAME_FIXES: dict[tuple[str, str], str] = {
    # ── Önceki oturumlardan ─────────────────────────────────────────────────
    ("omer selman karakose",    "14"): "Ömer Selman Karaköse",
    ("omer selman karakos1e4",  "14"): "Ömer Selman Karaköse",   # OCR bozuk
    ("b. altunzincir",          "14"): "Başak İrem Altunzincir",  # kısaltma
    ("basak irem altunzincir",  "14"): "Başak İrem Altunzincir",
    ("basar irem altunzincir",  "14"): "Başak İrem Altunzincir",  # yazım hatası
    ("c. colakogullari",        "10"): "Çağdaş Çolakoğulları",   # kısaltma
    ("cagdas colakogullari",    "10"): "Çağdaş Çolakoğulları",
    ("fatma berra ozer",        "13"): "Fatma Berra Özer",
    ("ahmet tuna atci",         "13"): "Ahmet Tuna Atcı",
    ("yagmur kont",             "13"): "Yağmur Kont",

    # ── İ / I / ı karışıklığı — giriş listesi & canlı kaynak düzeltmeleri ──
    ("ismail engin akdogan",    "12"): "İsmail Engin Akdoğan",
    ("kumsal ikra yasar",       "13"): "Kumsal İkra Yaşar",
    ("irmak akcengiz",          "14"): "Irmak AKCENGİZ",
    ("irmak barutcuoglu",       "13"): "Irmak BARUTÇUOĞLU",
    ("ilgaz firtina",           "12"): "Ilgaz FIRTINA",
    ("ilgim ertas",             "14"): "Ilgım ERTAŞ",
    ("inci tanriverdi",         "14"): "İnci TANRIVERDİ",
    ("melis ece ilter",         "14"): "Melis Ece İLTER",
    ("idil nilay karsli",       "14"): "İdil Nilay Karslı",
    ("berhan ilisik",           "10"): "Berhan İLİŞİK",
    ("eymen batu ibolar",       "10"): "Eymen Batu İBOLAR",
    ("ismail esad suslu",       "10"): "İsmail Esad SÜSLÜ",
    ("idil gulcan",             "11"): "İdil GÜLCAN",
    ("ilkut girayhan akyuz",    "12"): "İlkut Girayhan AKYÜZ",
    ("pars ikikardaslar",       "12"): "Pars İKİKARDAŞLAR",
    ("cemre ince",              "11"): "Cemre İNCE",
    ("ipek sozer",              "11"): "İpek SÖZER",
    ("ibrahim eren atakan",     "11"): "İbrahim Eren Atakan",
    ("ibrahim mutlu",           "13"): "İbrahim MUTLU",
    ("ikra sivaci",             "10"): "Ikra SIVACI",
    ("ipek gokce demirbasak",   "12"): "İpek Gökçe DEMİRBAŞAK",
    ("aras ipek",               "12"): "Aras İPEK",
    ("dogu ipek",               "14"): "Doğu İPEK",
    ("ismail otlak",            "13"): "İsmail OTLAK",
    ("elif inat",               "14"): "Elif İNAT",
    ("elif ipek bayrak",        "12"): "Elif İpek BAYRAK",
    ("nisa nur ilyan",          "14"): "Nisa Nur İLYAN",
}


def fix_city(c: str) -> str:
    return CITY_FIXES.get(c.lower().strip(), c.strip())


def norm(s: str) -> str:
    if not s:
        return ""
    return str(s).translate(TR_MAP).lower().strip()


def to_sec(t) -> float | None:
    if not t:
        return None
    s = str(t).strip().upper()
    if s in ("NT", "-", "DNS", "DQ", ""):
        return None
    try:
        if ":" in s:
            parts = s.split(":")
            return int(parts[0]) * 60 + float(parts[1])
        return float(s)
    except Exception:
        return None


def load_entry_list(path: str = ENTRY_PATH) -> list[dict]:
    """
    Giriş listesini yükler.
    Dönüş: her sporcu için dict listesi:
      {name, city, club, gender, yb, events: [{event, time_raw, time_sec}]}
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    # "SEM Kontrol Listesi" sheet varsa onu kullan
    ws = wb["SEM Kontrol Listesi"] if "SEM Kontrol Listesi" in wb.sheetnames else wb.active

    swimmers = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 17)]  # 1-16 (P = Kulüp)
        if not any(vals):
            continue

        city   = vals[1] or ""
        gender = vals[2] or ""   # Bayan / Erkek
        name   = vals[3] or ""
        yb     = str(vals[4]).strip() if vals[4] is not None else ""
        kulup  = vals[15] or ""  # P sütunu (indeks 15)
        if not name or not yb:
            continue

        # Şehir ve kulüp düzelt
        city_str  = fix_city(str(city))
        club_str  = str(kulup).strip() if kulup else ""

        # Sporcu adı düzelt
        name_str = str(name).strip()
        fixed = NAME_FIXES.get((norm(name_str), yb))
        if fixed:
            name_str = fixed

        # gender normalize → Kadın/Erkek
        gender_norm = "Erkek" if norm(gender) == "erkek" else "Kadın"

        events = []
        for i in range(4):
            ev   = vals[6 + i * 2]
            time = vals[7 + i * 2]
            if not ev:
                continue
            ev_str = str(ev).strip()
            if ev_str not in KNOWN_EVENTS:
                continue
            if yb in NO_50M_GROUPS and ev_str.startswith("50m"):
                continue
            t_sec = to_sec(time)
            events.append({
                "event":    ev_str,
                "time_raw": str(time).strip() if time else "NT",
                "time_sec": t_sec,
            })

        swimmers.append({
            "name":   name_str,
            "city":   city_str,
            "club":   club_str,
            "gender": gender_norm,
            "yb":     yb,
            "events": events,
        })

    wb.close()
    return swimmers


if __name__ == "__main__":
    swimmers = load_entry_list()
    print(f"Toplam sporcu: {len(swimmers)}")
    from collections import Counter
    grp = Counter((s["yb"], s["gender"]) for s in swimmers)
    for k, v in sorted(grp.items()):
        print(f"  {k}: {v}")
    clubs = Counter(s["club"] for s in swimmers if s["club"])
    print(f"\nKulup sayisi: {len(clubs)}")
    for c, n in clubs.most_common(5):
        print(f"  {n:3d}  {c}")
