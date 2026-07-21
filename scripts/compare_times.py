"""
SEM giriş süreleri vs sporcu_yildizlar Manual sheet karsılastırması.

Çalıştır:
    python scripts/compare_times.py

Çıktı:
    data/sure_guncelleme.xlsx  — Onay için hazır tablo
"""
import sys, os, re, unicodedata
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SEM_ENTRY   = r"C:\Users\Gokay\Desktop\Claude\2026 Temmuz SEM Final\SEM_kontrol_liste.xlsx"
MANUAL_XL   = r"C:\Users\Gokay\Desktop\Claude\Bölge Karmaları 2026\Çıktılar\Yıldızlar Yarış Sonuçları Çıktı\sporcu_yildizlar.xlsx"
OUT_XL      = r"C:\Users\Gokay\Desktop\Claude\2026 Temmuz SEM Final\data\sure_guncelleme.xlsx"

# Manual sheet kolonları → branş adı (SEM ile eşleşmesi için)
MANUAL_COLS = {
    6:  "50m Serbest",
    7:  "100m Serbest",
    8:  "200m Serbest",
    9:  "400m Serbest",
    10: "800m Serbest",
    11: "1500m Serbest",
    12: "50m Sırtüstü",
    13: "100m Sırtüstü",
    14: "200m Sırtüstü",
    15: "50m Kurbağalama",
    16: "100m Kurbağalama",
    17: "200m Kurbağalama",
    18: "50m Kelebek",
    19: "100m Kelebek",
    20: "200m Kelebek",
    21: "200m Karışık",
    22: "400m Karışık",
}

def normalize(name: str) -> str:
    """Büyük harf, Türkçe → ASCII, fazla boşluk sil."""
    if not name:
        return ""
    name = name.upper().strip()
    tr = str.maketrans("ÇĞİIÖŞÜçğışöşü", "CGIIOSUcgiossu")
    name = name.translate(tr)
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = re.sub(r"\s+", " ", name)
    return name

def to_sec(t: str) -> float | None:
    """
    Süresi saniyeye çevir.
    Desteklenen formatlar: M:SS.cc  MM:SS.cc  H:MM:SS.cc
    """
    if not t:
        return None
    t = str(t).strip().replace(",", ".")
    parts = t.split(":")
    try:
        if len(parts) == 2:
            return int(parts[0]) * 60 + float(parts[1])
        elif len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    except (ValueError, IndexError):
        return None
    return None

def fmt(sec: float | None) -> str:
    """Saniyeyi M:SS.cc formatına çevir."""
    if sec is None:
        return ""
    m = int(sec) // 60
    s = sec - m * 60
    return f"{m}:{s:05.2f}"

# ─── 1. SEM giriş listesi ────────────────────────────────────────────────────
wb_sem = openpyxl.load_workbook(SEM_ENTRY, data_only=True)
ws_sem = wb_sem["SEM Kontrol Listesi"]

sem_athletes = []  # [{name, yb, city, club, gender, events:{branş:süre_str}}, ...]
for r in range(2, ws_sem.max_row + 1):
    name = ws_sem.cell(r, 4).value
    if not name:
        continue
    yb     = str(ws_sem.cell(r, 5).value or "").strip()
    city   = str(ws_sem.cell(r, 2).value or "").strip()
    gender = str(ws_sem.cell(r, 3).value or "").strip()
    club   = str(ws_sem.cell(r, 16).value or "").strip()
    events = {}
    for i in range(4):
        brans = ws_sem.cell(r, 7 + i*2).value
        sure  = ws_sem.cell(r, 8 + i*2).value
        if brans and sure:
            events[str(brans).strip()] = str(sure).strip()
    sem_athletes.append({
        "name": str(name).strip(), "yb": yb, "city": city,
        "gender": gender, "club": club, "events": events, "row": r
    })

# ─── 2. Manual sheet ─────────────────────────────────────────────────────────
wb_manual = openpyxl.load_workbook(MANUAL_XL, data_only=True)
ws_m = wb_manual["Manual"]

# {norm_name: {branş: en_iyi_süre_str}}
manual_db: dict[str, dict[str, str]] = {}
for r in range(2, ws_m.max_row + 1):
    isim = ws_m.cell(r, 1).value
    if not isim:
        continue
    key = normalize(str(isim))
    if key not in manual_db:
        manual_db[key] = {}
    for col, brans in MANUAL_COLS.items():
        val = ws_m.cell(r, col).value
        if not val:
            continue
        s_new = to_sec(str(val))
        if s_new is None:
            continue
        if brans not in manual_db[key]:
            manual_db[key][brans] = str(val).strip()
        else:
            s_old = to_sec(manual_db[key][brans])
            if s_old is None or s_new < s_old:
                manual_db[key][brans] = str(val).strip()

# ─── 3. Karşılaştır ──────────────────────────────────────────────────────────
results   = []   # iyileşme var olanlar
no_match  = []   # eşleşemeyen
same      = []   # eşleşti ama değişiklik yok

for ath in sem_athletes:
    key = normalize(ath["name"])
    if key not in manual_db:
        no_match.append(ath)
        continue

    m_times = manual_db[key]
    improved = False
    for brans, sem_sure in ath["events"].items():
        sem_sec = to_sec(sem_sure)
        if sem_sec is None:
            continue
        if brans not in m_times:
            continue
        man_sec = to_sec(m_times[brans])
        if man_sec is None:
            continue
        fark = sem_sec - man_sec  # pozitif → sporcu daha iyi (daha hızlı)
        if man_sec < sem_sec:
            improved = True
            results.append({
                "name":    ath["name"],
                "yb":      ath["yb"],
                "city":    ath["city"],
                "club":    ath["club"],
                "gender":  ath["gender"],
                "brans":   brans,
                "sem":     sem_sure,
                "manual":  m_times[brans],
                "fark_s":  fark,
                "row":     ath["row"],
            })

# ─── 4. Excel çıktısı ─────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUT_XL), exist_ok=True)
wb_out = openpyxl.Workbook()

# ── Sheet 1: Güncellenecekler
ws_out = wb_out.active
ws_out.title = "Güncellenecekler"

HDR_FILL   = PatternFill("solid", fgColor="1E3A5F")
GREEN_FILL = PatternFill("solid", fgColor="0F4C25")
RED_FILL   = PatternFill("solid", fgColor="4C0F0F")
hdr_font   = Font(bold=True, color="FFFFFF", size=11)
thin = Border(
    left=Side(style="thin", color="404040"),
    right=Side(style="thin", color="404040"),
    top=Side(style="thin", color="404040"),
    bottom=Side(style="thin", color="404040"),
)

headers_out = ["#", "Ad Soyad", "YB", "Şehir", "Cinsiyet", "Kulüp",
               "Branş", "SEM Süresi", "Yeni Süre", "Fark (s)", "Onayla (E/H)"]
ws_out.append(headers_out)
for c, h in enumerate(headers_out, 1):
    cell = ws_out.cell(1, c)
    cell.fill = HDR_FILL
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin

col_widths = [4, 28, 5, 12, 9, 32, 22, 12, 12, 10, 14]
for c, w in enumerate(col_widths, 1):
    ws_out.column_dimensions[get_column_letter(c)].width = w

for i, r in enumerate(sorted(results, key=lambda x: (-x["fark_s"], x["name"])), 1):
    row_data = [
        i, r["name"], r["yb"], r["city"], r["gender"], r["club"],
        r["brans"], r["sem"], r["manual"], f"+{r['fark_s']:.2f}s", "S"
    ]
    ws_out.append(row_data)
    rn = i + 1
    # Renk: fazla fark → daha yeşil
    fill = PatternFill("solid", fgColor="0D5C2B") if r["fark_s"] >= 3 else PatternFill("solid", fgColor="1A3A20")
    for c in range(1, 12):
        cell = ws_out.cell(rn, c)
        cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center" if c in (1,3,5,10,11) else "left")

# "Onayla" sütunu = sarı arkaplan
yellow = PatternFill("solid", fgColor="3A3000")
for rn in range(2, len(results) + 2):
    ws_out.cell(rn, 11).fill = yellow
    ws_out.cell(rn, 11).font = Font(bold=True, color="FFD700")

ws_out.freeze_panes = "A2"

# ── Sheet 2: Eşleşmeyenler
ws2 = wb_out.create_sheet("Eslesmeyenler")
ws2.append(["Ad Soyad", "YB", "Şehir", "Branş Sayısı"])
ws2.cell(1, 1).fill = HDR_FILL; ws2.cell(1, 1).font = hdr_font
ws2.cell(1, 2).fill = HDR_FILL; ws2.cell(1, 2).font = hdr_font
ws2.cell(1, 3).fill = HDR_FILL; ws2.cell(1, 3).font = hdr_font
ws2.cell(1, 4).fill = HDR_FILL; ws2.cell(1, 4).font = hdr_font
for ath in sorted(no_match, key=lambda x: x["name"]):
    ws2.append([ath["name"], ath["yb"], ath["city"], len(ath["events"])])
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["C"].width = 14

wb_out.save(OUT_XL)

# ─── 5. Özet ──────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"  SEM Giriş Listesi    : {len(sem_athletes)} sporcu")
print(f"  Manual Sheet         : {len(manual_db)} sporcu (unique isim)")
print(f"  Eşleşen              : {len(sem_athletes) - len(no_match)} sporcu")
print(f"  Eşleşmeyen           : {len(no_match)} sporcu")
print(f"  Güncellenecek branş  : {len(results)} adet")
print(f"{'='*60}")
print(f"\nGüncelleme adayları ({len(results)} branş):")
prev = ""
for r in sorted(results, key=lambda x: (-x["fark_s"], x["name"])):
    lbl = r["name"] if r["name"] != prev else "  └─"
    prev = r["name"]
    print(f"  {lbl:30s}  YB={r['yb']}  {r['brans']:20s}  SEM={r['sem']:10s}  Yeni={r['manual']:10s}  +{r['fark_s']:.2f}s")

print(f"\nOnayland kelimesi için:")
print(f"  {OUT_XL}")
print(f"\nOnayladıktan sonra çalıştır:")
print(f"  python scripts/apply_updates.py")
