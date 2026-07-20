"""
SEM_kontrol_liste.xlsx icindeki sehir ve sporcu adi duzeltmeleri.
Kullanim: python scripts/fix_excel_names.py
"""
import sys
import os
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl

_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _dir)
from sem_entry import CITY_FIXES, NAME_FIXES, fix_city, norm
from sem_config import ENTRY_PATH


def main():
    wb = openpyxl.load_workbook(ENTRY_PATH)
    ws = wb.active

    city_fixes_applied = []
    name_fixes_applied = []

    for r in range(2, ws.max_row + 1):
        # Sutun B (2) = Sehir
        city_cell = ws.cell(r, 2)
        if city_cell.value:
            raw = str(city_cell.value).strip()
            fixed = fix_city(raw)
            if raw != fixed:
                city_fixes_applied.append(f"  Satir {r}: '{raw}' -> '{fixed}'")
                city_cell.value = fixed

        # Sutun D (4) = Ad Soyad, Sutun E (5) = YB
        name_cell = ws.cell(r, 4)
        yb_cell   = ws.cell(r, 5)
        if name_cell.value and yb_cell.value is not None:
            raw_name = str(name_cell.value).strip()
            yb       = str(yb_cell.value).strip()
            key      = (norm(raw_name), yb)
            fixed    = NAME_FIXES.get(key)
            if fixed and raw_name != fixed:
                name_fixes_applied.append(f"  Satir {r}: '{raw_name}' -> '{fixed}'")
                name_cell.value = fixed

    wb.save(ENTRY_PATH)
    print("Excel kaydedildi:", ENTRY_PATH)
    print(f"\nSehir duzeltmeleri ({len(city_fixes_applied)}):")
    for f in city_fixes_applied:
        print(f)
    print(f"\nAd duzeltmeleri ({len(name_fixes_applied)}):")
    for f in name_fixes_applied:
        print(f)
    if not city_fixes_applied and not name_fixes_applied:
        print("  (Duzeltme yapilmadi - zaten dogru veya eslesen satir bulunamadi)")


if __name__ == "__main__":
    main()
